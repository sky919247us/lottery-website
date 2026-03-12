"""
從 Excel 匯入經銷商資料到 SQLite
讀取「前次台彩」與「前次運彩」工作表
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from app.model.database import SessionLocal, init_db
from app.model.retailer import Retailer


EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "..",
    "台灣彩券 台灣運彩異動明細.xlsx",
)

# 台彩 / 運彩工作表名稱與對應來源標籤
SHEETS = [
    ("前次台彩", "台灣彩券"),
    ("前次運彩", "台灣運彩"),
]


def normalize_city(city: str) -> str:
    """統一縣市名稱（臺→台）"""
    if not city:
        return ""
    return city.replace("臺", "台")


def import_retailers():
    """匯入經銷商資料"""
    init_db()
    db = SessionLocal()

    # 清除舊資料
    deleted = db.query(Retailer).delete()
    db.commit()
    print(f"已清除 {deleted} 筆舊資料")

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    total_imported = 0

    for sheet_name, source_label in SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"找不到工作表 [{sheet_name}]，跳過")
            continue

        ws = wb[sheet_name]
        count = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 欄位：縣市, 行政區, 地址, 投注站名稱, Google 地圖連結
            if len(row) < 4:
                continue

            city_raw = str(row[0] or "").strip()
            district = str(row[1] or "").strip()
            address = str(row[2] or "").strip()
            name = str(row[3] or "").strip()

            # 跳過空白列
            if not name or not address:
                continue

            city = normalize_city(city_raw)

            retailer = Retailer(
                name=name,
                address=address,
                city=city,
                district=district,
                source=source_label,
                lat=None,
                lng=None,
                isActive=True,
            )
            db.add(retailer)
            count += 1

            # 每 500 筆 commit 一次
            if count % 500 == 0:
                db.commit()
                print(f"  [{sheet_name}] 已匯入 {count} 筆...")

        db.commit()
        total_imported += count
        print(f"[{sheet_name}] 匯入完成：{count} 筆")

    db.close()
    wb.close()
    print(f"\n全部匯入完成！共 {total_imported} 筆經銷商資料")


if __name__ == "__main__":
    import_retailers()
